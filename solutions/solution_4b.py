# Instructor reference solution — Exercise 4B
import pandas as pd
from openpyxl.styles import PatternFill

df = pd.read_csv('../data/transactions_clean.csv')

# 1. Total and average amount by segment
summary = df.groupby('segment')['amount'].agg(total_amount='sum', avg_amount='mean').reset_index()

# 2. Flag high-value transactions
df['high_value'] = df['amount'] > 10000

# 3. Export summary to CSV
summary.to_csv('summary_report.csv', index=False)

# --- Extension ---
flagged = df[df['high_value']]
branch_breakdown = (
    df[df['high_value']].groupby('branch').size().reset_index(name='high_value_count')
)

with pd.ExcelWriter('transaction_report.xlsx', engine='openpyxl') as writer:
    summary.to_excel(writer, sheet_name='Segment Summary', index=False)
    flagged.to_excel(writer, sheet_name='Flagged Transactions', index=False)
    branch_breakdown.to_excel(writer, sheet_name='Branch Breakdown', index=False)

    # Advanced bonus: red fill on flagged rows in 'Flagged Transactions'
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    sheet = writer.sheets['Flagged Transactions']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.fill = red_fill

print(f"Segment summary:\n{summary}\n")
print(f"Branch breakdown:\n{branch_breakdown}")
